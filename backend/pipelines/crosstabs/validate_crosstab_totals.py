# -*- coding: utf-8 -*-
"""
STAGE 4 -- CROSSTAB TOTAL VALIDATION
====================================
Runs after the crosstab stage. Opens EVERY exported workbook, finds EVERY
crosstab table inside it, and checks that the printed counts actually add up to
the printed Total -- in BOTH directions:

    COLUMN-WISE : for each column, sum(data rows)    == the 'Total' row
    ROW-WISE    : for each data row, sum(data cols)  == the 'Total' column

Every cell is formatted "<count> (<pct>%)"; only the count is checked.

Writes a report workbook with three sheets:
    'Run summary'  - headline numbers and the overall verdict
    'Table totals' - one line per table: its totals and its check results
    'Mismatches'   - one line per failing check, with a detailed reason

This module only READS the outputs and WRITES the report. It never changes a
crosstab, and a failure here cannot corrupt pipeline results.
"""
import os
import re
from collections import Counter

import pandas as pd

COUNT_RE = re.compile(r'^\s*(-?\d+)\s*(?:\(|$)')

# ----------------------------------------------------------------- reason text
REASON_EXCL_NORMAL = (
    "BY DESIGN - not an error. This is an '(excluding Normal)' twin table produced by the "
    "GENERATE_ANTHRO_STATUS_CT add-on. The 'Normal' category row is removed from DISPLAY ONLY; the "
    "counts, the percentages and the Total row all stay computed on the FULL N *including* Normal. "
    "That is deliberate - these tables express each malnutrition category as a share of ALL children, "
    "not as a share of non-Normal children. If the visible rows DID add up to the Total, the "
    "percentages would be wrong. Compare any such table with its full twin directly above it: the "
    "percentages and the Total row are identical, only the Normal row is hidden. "
    "No action required."
)

REASON_BLOCK_ZNA = (
    "HIDDEN 'no block' COLUMN. The crosstab generator computes the Total with margins=True FIRST and "
    "only afterwards drops the 'z_NA' row/column from display. Cases whose block could not be "
    "determined therefore sit inside the Total without appearing under any block column, so the "
    "visible block columns add up to less than the Total. Root cause is missing source data: "
    "'Phc Taluka_CR' is blank in case registration for these learners, and neither the MASD combined "
    "sheet nor the DOCS sheet could supply a block for them. {live} "
    "The affected learners and cases are listed in no_block_case_list_<code>_<V>.xlsx. "
    "ACTION: fill Block/Taluk for these learners at source (or in MASD), or set "
    "EXCLUDE_UNDEFINED_BLOCK = True to drop them from the analysis, after which every block table "
    "balances exactly."
)

REASON_ROLE_ZNA = (
    "HIDDEN 'z_NA' CATEGORY ROW. The Total was computed before the z_NA row was suppressed from "
    "display, so the visible rows fall short of the Total. Cases with no role/role group defined are "
    "normally removed from the analysis by EXCLUDE_UNDEFINED_ROLE_GROUP = True, so this should not "
    "normally appear - if it does, check that switch and the Role and department sheet."
)

REASON_UNKNOWN = (
    "UNEXPLAINED - INVESTIGATE. This mismatch does not match any known display-suppression rule "
    "(the '(excluding Normal)' twin tables, or a hidden z_NA row/column). Something else is wrong: "
    "check the source data for this variable, and check whether a category label was renamed in the "
    "Role and department sheet without the corresponding code/config being updated."
)


def cell_count(v):
    """Pull the leading integer out of a '123 (45.6%)' cell."""
    if isinstance(v, str):
        m = COUNT_RE.match(v)
        return int(m.group(1)) if m else None
    if isinstance(v, (int, float)) and not pd.isna(v):
        return int(v)
    return None


def find_tables(d):
    """Yield (title, header, data_row_idx, total_row_idx) for each table in a sheet."""
    nrow = d.shape[0]
    i = 0
    while i < nrow:
        t = d.iat[i, 0]
        if isinstance(t, str) and ' vs ' in t:
            title = t.strip()
            h = i + 1
            while h < nrow:
                if any(isinstance(d.iat[h, c], str) and str(d.iat[h, c]).strip()
                       for c in range(1, d.shape[1])):
                    break
                h += 1
            if h >= nrow:
                i += 1
                continue
            header = [(c, str(d.iat[h, c]).strip()) for c in range(1, d.shape[1])
                      if isinstance(d.iat[h, c], str) and str(d.iat[h, c]).strip()]
            rows, total_row = [], None
            r = h + 1
            while r < nrow:
                lab = d.iat[r, 0]
                if isinstance(lab, str) and lab.strip():
                    if lab.strip().lower() == 'total':
                        total_row = r
                        break
                    if ' vs ' in lab:
                        break
                    rows.append(r)
                r += 1
            if total_row is not None and rows and header:
                yield title, header, rows, total_row
            i = (total_row if total_row is not None else h) + 1
        else:
            i += 1


def _live_block_note(outputs_dir):
    """Read the no-block list so the reason text carries this run's real numbers."""
    try:
        for root, _, files in os.walk(outputs_dir):
            for f in files:
                if f.startswith('no_block_case_list') and f.endswith('.xlsx') \
                        and not f.startswith('~$'):
                    x = pd.ExcelFile(os.path.join(root, f))
                    lr = x.parse('No-block learners')
                    cs = x.parse('No-block cases')
                    return (f"In THIS run {len(cs)} case row(s) across {len(lr)} learner(s) "
                            f"still have no block.")
    except Exception:
        pass
    return ""


def classify(title, direction, diff, live_note):
    """Return (category, detailed reason) for one failing check."""
    t = title.lower()
    if direction == 'Column-wise' and 'excluding normal' in t:
        return 'By design (excluding Normal)', REASON_EXCL_NORMAL
    if direction == 'Row-wise' and t.endswith('vs blocks'):
        return 'Missing source data (no block)', REASON_BLOCK_ZNA.format(live=live_note).strip()
    if direction == 'Column-wise' and t.startswith('blocks vs'):
        return 'Missing source data (no block)', REASON_BLOCK_ZNA.format(live=live_note).strip()
    if 'role group' in t or 'department' in t:
        return 'Hidden z_NA category', REASON_ROLE_ZNA
    return 'UNEXPLAINED', REASON_UNKNOWN


def validate(outputs_dir, report_path):
    outputs_dir = str(outputs_dir)
    live_note = _live_block_note(outputs_dir)

    table_rows, mismatch_rows = [], []
    skipped, n_files = [], 0

    paths = []
    for root, _, files in os.walk(outputs_dir):
        for f in sorted(files):
            if f.endswith('.xlsx') and not f.startswith('~$') \
                    and os.path.basename(report_path) != f:
                paths.append(os.path.join(root, f))

    for p in paths:
        rel = os.path.relpath(p, outputs_dir)
        folder = os.path.dirname(rel) or '.'
        try:
            x = pd.ExcelFile(p)
        except Exception as e:
            skipped.append((rel, f'{type(e).__name__}: {e}'))
            continue
        n_files += 1
        found_any = False
        for sh in x.sheet_names:
            try:
                d = x.parse(sh, header=None)
            except Exception as e:
                skipped.append((f'{rel}[{sh}]', f'{type(e).__name__}: {e}'))
                continue
            if d.empty:
                continue
            for title, header, rows, total_row in find_tables(d):
                found_any = True
                cols = [c for c, lab in header if lab.lower() != 'total']
                labels = {c: lab for c, lab in header}
                tcol = next((c for c, lab in header if lab.lower() == 'total'), None)

                c_n = c_ok = r_n = r_ok = 0

                # --------------------------------------------- column-wise
                for c in cols + ([tcol] if tcol is not None else []):
                    tot = cell_count(d.iat[total_row, c])
                    parts = [cell_count(d.iat[r, c]) for r in rows]
                    if tot is None or any(v is None for v in parts):
                        continue
                    c_n += 1
                    s = sum(parts)
                    if s == tot:
                        c_ok += 1
                    else:
                        cat, why = classify(title, 'Column-wise', s - tot, live_note)
                        mismatch_rows.append({
                            'File': os.path.basename(p), 'Folder': folder, 'Sheet': sh,
                            'Table': title, 'Direction': 'Column-wise',
                            'Line checked': f'column "{labels.get(c, c)}"',
                            'Sum of visible cells': s, 'Printed Total': tot,
                            'Difference': s - tot, 'Category': cat, 'Detailed reason': why,
                        })

                # ------------------------------------------------ row-wise
                if tcol is not None:
                    for r in rows + [total_row]:
                        tot = cell_count(d.iat[r, tcol])
                        parts = [cell_count(d.iat[r, c]) for c in cols]
                        if tot is None or any(v is None for v in parts):
                            continue
                        r_n += 1
                        s = sum(parts)
                        if s == tot:
                            r_ok += 1
                        else:
                            cat, why = classify(title, 'Row-wise', s - tot, live_note)
                            lab = str(d.iat[r, 0]).strip()
                            mismatch_rows.append({
                                'File': os.path.basename(p), 'Folder': folder, 'Sheet': sh,
                                'Table': title, 'Direction': 'Row-wise',
                                'Line checked': f'row "{lab}"',
                                'Sum of visible cells': s, 'Printed Total': tot,
                                'Difference': s - tot, 'Category': cat, 'Detailed reason': why,
                            })

                grand = cell_count(d.iat[total_row, tcol]) if tcol is not None else None
                rv, _, cv = title.partition(' vs ')
                table_rows.append({
                    'File': os.path.basename(p), 'Folder': folder, 'Sheet': sh,
                    'Table': title, 'Row variable': rv.strip(), 'Column variable': cv.strip(),
                    'Data rows': len(rows), 'Data columns': len(cols),
                    'Grand Total': grand,
                    'Column-wise checks': c_n, 'Column-wise OK': c_ok,
                    'Column-wise mismatch': c_n - c_ok,
                    'Row-wise checks': r_n, 'Row-wise OK': r_ok,
                    'Row-wise mismatch': r_n - r_ok,
                    'Status': 'OK' if (c_n - c_ok) + (r_n - r_ok) == 0 else 'MISMATCH',
                })
        if not found_any:
            skipped.append((rel, 'no crosstab tables (list/export file)'))

    tdf = pd.DataFrame(table_rows)
    mdf = pd.DataFrame(mismatch_rows)

    cats = Counter(mdf['Category']) if len(mdf) else Counter()
    unexplained = int(cats.get('UNEXPLAINED', 0))
    summary = [
        ('Workbooks scanned', n_files),
        ('Crosstab tables checked', len(tdf)),
        ('Total individual checks', int(tdf[['Column-wise checks', 'Row-wise checks']].sum().sum()) if len(tdf) else 0),
        ('Column-wise checks', int(tdf['Column-wise checks'].sum()) if len(tdf) else 0),
        ('Column-wise OK', int(tdf['Column-wise OK'].sum()) if len(tdf) else 0),
        ('Column-wise MISMATCH', int(tdf['Column-wise mismatch'].sum()) if len(tdf) else 0),
        ('Row-wise checks', int(tdf['Row-wise checks'].sum()) if len(tdf) else 0),
        ('Row-wise OK', int(tdf['Row-wise OK'].sum()) if len(tdf) else 0),
        ('Row-wise MISMATCH', int(tdf['Row-wise mismatch'].sum()) if len(tdf) else 0),
        ('', ''),
        ('Tables fully OK', int((tdf['Status'] == 'OK').sum()) if len(tdf) else 0),
        ('Tables with a mismatch', int((tdf['Status'] == 'MISMATCH').sum()) if len(tdf) else 0),
        ('', ''),
    ]
    for c, n in cats.most_common():
        summary.append((f'Mismatches - {c}', n))
    summary += [
        ('', ''),
        ('UNEXPLAINED mismatches', unexplained),
        ('VERDICT', 'PASS - every mismatch is explained' if unexplained == 0
         else f'FAIL - {unexplained} unexplained mismatch(es), see Mismatches sheet'),
    ]
    sdf = pd.DataFrame(summary, columns=['Measure', 'Value'])

    if len(mdf) == 0:
        mdf = pd.DataFrame([{'File': '', 'Folder': '', 'Sheet': '', 'Table': '',
                             'Direction': '', 'Line checked': 'NO MISMATCHES FOUND',
                             'Sum of visible cells': '', 'Printed Total': '',
                             'Difference': '', 'Category': '', 'Detailed reason': ''}])

    with pd.ExcelWriter(report_path, engine='openpyxl') as w:
        sdf.to_excel(w, sheet_name='Run summary', index=False)
        tdf.to_excel(w, sheet_name='Table totals', index=False)
        mdf.to_excel(w, sheet_name='Mismatches', index=False)
        if skipped:
            pd.DataFrame(skipped, columns=['File', 'Why not checked']).to_excel(
                w, sheet_name='Not checked', index=False)
        _format(w, sdf, tdf, mdf)

    return {'files': n_files, 'tables': len(tdf),
            'col_mismatch': int(tdf['Column-wise mismatch'].sum()) if len(tdf) else 0,
            'row_mismatch': int(tdf['Row-wise mismatch'].sum()) if len(tdf) else 0,
            'unexplained': unexplained, 'categories': dict(cats),
            'report': str(report_path)}


def _format(writer, sdf, tdf, mdf):
    from openpyxl.styles import Alignment, Font, PatternFill
    hdr = PatternFill('solid', fgColor='1F4E78')
    for name, widths in [
        ('Run summary', {'A': 34, 'B': 62}),
        ('Table totals', {'A': 52, 'B': 12, 'C': 16, 'D': 44, 'E': 26, 'F': 22,
                          'G': 10, 'H': 12, 'I': 12, 'J': 15, 'K': 14, 'L': 16,
                          'M': 13, 'N': 12, 'O': 14, 'P': 11}),
        ('Mismatches', {'A': 52, 'B': 12, 'C': 16, 'D': 44, 'E': 13, 'F': 26,
                        'G': 15, 'H': 13, 'I': 11, 'J': 30, 'K': 120}),
    ]:
        ws = writer.sheets[name]
        for col, wd in widths.items():
            ws.column_dimensions[col].width = wd
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.freeze_panes = 'A2'

    red = PatternFill('solid', fgColor='FFC7CE')
    amber = PatternFill('solid', fgColor='FFF2CC')
    green = PatternFill('solid', fgColor='E2EFDA')

    ws = writer.sheets['Table totals']
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=16).fill = green if ws.cell(row=r, column=16).value == 'OK' else amber

    ws = writer.sheets['Mismatches']
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=11).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = 92
        if str(ws.cell(row=r, column=10).value) == 'UNEXPLAINED':
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = red
        elif str(ws.cell(row=r, column=10).value).startswith('By design'):
            ws.cell(row=r, column=10).fill = green
        else:
            ws.cell(row=r, column=10).fill = amber

    ws = writer.sheets['Run summary']
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = Alignment(vertical='top', wrap_text=True)
        if str(ws.cell(row=r, column=1).value) == 'VERDICT':
            ok = str(ws.cell(row=r, column=2).value).startswith('PASS')
            for c in (1, 2):
                ws.cell(row=r, column=c).fill = green if ok else red
                ws.cell(row=r, column=c).font = Font(bold=True)


if __name__ == '__main__':
    import sys
    od = sys.argv[1] if len(sys.argv) > 1 else 'outputs'
    rp = sys.argv[2] if len(sys.argv) > 2 else 'crosstab_total_validation_report.xlsx'
    print(validate(od, rp))

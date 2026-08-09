"""Growth-monitoring (LAP) chart endpoints.

Derives per-child visit series from submitted form responses: a *visit* is the
set of responses sharing (child, assessment_date). Weight/length come from the
``growth_monitoring`` flat form; the breastfeeding / complementary_feeding
responses filed the same day determine the visit's "sources" combination, which
the charts color-code. Percentile backgrounds come from the WHO standards
tables (see app/who_growth.py).

Endpoints:
- GET /api/growth/standards            → WHO percentile curves (public reference data)
- GET /api/growth/my-cases             → the logged-in learner's cases
- GET /api/admin/growth/monitor        → all learner-mother-child cases (admin, ?district=)
- GET /api/admin/growth/responses/{id} → response detail for the visit modal (admin;
                                         learners use GET /api/forms/responses/{id})

Admin endpoints live under /api/admin/* because the frontend axios client
attaches the admin JWT only to that prefix.
"""

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app import growth_summary_mock as gsm  # MOCK: summary-table demo values (removable)
from app import models, projects
from app.database import get_db
from app.dependencies import get_current_admin, get_verified_user
from app.routers.forms import PROTEIN_HIGH_TOTAL_G, _serialize_detail
from app.who_growth import percentile_curves, zscore_for_value

router = APIRouter(prefix="/api/growth", tags=["growth"])
admin_router = APIRouter(prefix="/api/admin/growth", tags=["growth"])

GROWTH_FORM_KEY = "growth_monitoring"
VISIT_FORM_KEYS = ("growth_monitoring", "breastfeeding", "complementary_feeding")

# Plausible measurement ranges (kg / cm) — reject a mis-matched number field
# (the admin can rename/re-create fields, so extraction is heuristic).
WEIGHT_RANGE = (0.5, 35.0)
LENGTH_RANGE = (25.0, 130.0)


# ── Visit-series assembly ────────────────────────────────────────────────────

def _extract_metrics(answers_json: Any) -> Dict[str, Optional[float]]:
    """Pull the weight/length measurements out of a growth response's answer
    snapshots. The growth form is admin-editable, so fields are matched by
    id/label (contains "weight" / "length"|"height") with a range sanity check
    rather than by a hardcoded field id."""
    metrics: Dict[str, Optional[float]] = {"weight": None, "length": None}
    for snap in answers_json or []:
        if not isinstance(snap, dict) or not snap.get("value"):
            continue
        if snap.get("questionType") not in (None, "number"):
            continue
        key = f"{snap.get('nodeId') or ''} {snap.get('question') or ''}".lower()
        try:
            value = float(snap["value"])
        except (TypeError, ValueError):
            continue
        if "weight" in key and metrics["weight"] is None:
            if WEIGHT_RANGE[0] <= value <= WEIGHT_RANGE[1]:
                metrics["weight"] = value
        elif ("length" in key or "height" in key) and metrics["length"] is None:
            if LENGTH_RANGE[0] <= value <= LENGTH_RANGE[1]:
                metrics["length"] = value
    return metrics


def _build_cases(
    db: Session,
    *,
    user_id: Optional[int] = None,
    district_slug: Optional[str] = None,
    role: Optional[str] = None,
    department: Optional[str] = None,
    learner_id: Optional[int] = None,
    child_id: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """One case per child: identity of the learner-mother-child triple plus the
    chronological visit series. Children without visits are included (they show
    as empty cases and let the UI offer the pair in filters).

    The learner is LEFT-joined: a mother's registered_by_user_id is nullable
    (ondelete=SET NULL), so a child whose registering account was removed still
    appears — as an orphaned case with learner=None — rather than silently
    vanishing from the admin monitor.
    """
    rows = (
        db.query(
            models.Child.id, models.Child.child_uid, models.Child.child_name,
            models.Child.gender, models.Child.dob, models.Child.birth_weight,
            models.Child.birth_length, models.Child.adoption_date,
            models.Mother.id, models.Mother.mother_uid, models.Mother.mother_name,
            models.User.id, models.User.full_name, models.User.email,
            models.User.role, models.User.department,
            models.ProgramDistrict.name,
        )
        .join(models.Mother, models.Child.mother_id == models.Mother.id)
        .outerjoin(models.User, models.Mother.registered_by_user_id == models.User.id)
        .outerjoin(models.ProgramDistrict, models.User.program_district_id == models.ProgramDistrict.id)
    )
    if user_id is not None:
        # scope by the FK column so my-cases semantics are unchanged
        rows = rows.filter(models.Mother.registered_by_user_id == user_id)
    if learner_id is not None:
        rows = rows.filter(models.User.id == learner_id)
    if child_id is not None:
        rows = rows.filter(models.Child.id == child_id)
    if role:
        rows = rows.filter(models.User.role == role)
    # department is resolved (mock-fillable) after the query, so it's filtered below.
    if district_slug:
        pd = (
            db.query(models.ProgramDistrict)
            .filter(models.ProgramDistrict.slug == district_slug)
            .first()
        )
        if not pd:
            return []
        # A state project covers its district projects' learners too.
        rows = rows.filter(models.User.program_district_id.in_(projects.member_project_ids(db, pd)))
    rows = rows.order_by(models.User.full_name, models.Mother.mother_name, models.Child.id).all()

    cases: Dict[int, Dict[str, Any]] = {}
    for (child_id, child_uid, child_name, gender, dob, birth_weight, birth_length, adoption_date,
         mother_id, mother_uid, mother_name,
         l_user_id, learner_name, learner_email, learner_role, learner_department,
         district_name) in rows:
        cases[child_id] = {
            "child": {
                "id": child_id,
                "uid": child_uid,
                "name": child_name,
                "gender": gender,
                "dob": dob.isoformat() if dob else None,
                "birth_weight": birth_weight,
                "birth_length": birth_length,
                "adoption_date": adoption_date.isoformat() if adoption_date else None,
            },
            "mother": {"id": mother_id, "uid": mother_uid, "name": mother_name},
            "learner": (
                {
                    "id": l_user_id,
                    "name": learner_name or learner_email,
                    "email": learner_email,
                    "role": learner_role,
                    "department": learner_department,
                    "district": district_name,
                }
                if l_user_id is not None
                else {"id": None, "name": None, "email": None,
                      "role": None, "department": None, "district": district_name}
            ),
            "visits": [],
        }

    # Department is mock-fillable (demo learners have none); resolve it here so the
    # value is identical across the table, the charts and the filter dropdown, then
    # apply the department filter on the resolved value.
    for case in cases.values():
        case["learner"]["department"] = gsm.resolved_department(  # MOCK: department fill
            case["learner"]["id"], case["learner"]["department"]
        )
    if department:
        cases = {cid: c for cid, c in cases.items() if c["learner"]["department"] == department}

    if not cases:
        return []

    child_ids = list(cases.keys())

    # Visit skeleton: which forms were filed per (child, date). Column-pruned —
    # the JSON answer blobs are NOT loaded here (they can be large, and BF/CF
    # blobs are never parsed for the chart).
    visit_rows = (
        db.query(
            models.FormResponse.child_id,
            models.FormResponse.assessment_date,
            models.FormResponse.form_key,
            models.FormResponse.id,
        )
        .filter(
            models.FormResponse.child_id.in_(child_ids),
            models.FormResponse.form_key.in_(VISIT_FORM_KEYS),
            models.FormResponse.status == "submitted",
        )
        .order_by(models.FormResponse.assessment_date, models.FormResponse.id)
        .all()
    )

    # Weight/length come only from growth responses, so load answers_json for
    # just those rows (one narrow query) instead of every visit form.
    metrics_by_response: Dict[int, Dict[str, Optional[float]]] = {}
    growth_ids = [rid for (_, _, fk, rid) in visit_rows if fk == GROWTH_FORM_KEY]
    if growth_ids:
        for rid, answers in (
            db.query(models.FormResponse.id, models.FormResponse.answers_json)
            .filter(models.FormResponse.id.in_(growth_ids))
            .all()
        ):
            metrics_by_response[rid] = _extract_metrics(answers)

    # visit key = (child_id, assessment_date)
    visits: Dict[tuple, Dict[str, Any]] = defaultdict(lambda: {"forms": {}, "weight": None, "length": None})
    for child_id, assessment_date, form_key, response_id in visit_rows:
        if not assessment_date:
            continue
        visit = visits[(child_id, assessment_date)]
        visit["forms"][form_key] = response_id
        if form_key == GROWTH_FORM_KEY:
            metrics = metrics_by_response.get(response_id, {"weight": None, "length": None})
            visit["weight"] = metrics["weight"]
            visit["length"] = metrics["length"]

    for (child_id, visit_date), visit in sorted(visits.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        case = cases.get(child_id)
        if case is None:
            continue
        dob_iso = case["child"]["dob"]
        age_days = None
        if dob_iso:
            age_days = (visit_date - date.fromisoformat(dob_iso)).days
        case["visits"].append({
            "date": visit_date.isoformat(),
            "age_days": age_days,
            "weight": visit["weight"],
            "length": visit["length"],
            "sources": sorted(visit["forms"].keys()),
            "forms": visit["forms"],
        })

    return list(cases.values())


# ── Learner-level summary table ──────────────────────────────────────────────
#
# One row per learner-mother-child case. Everything below is REAL except the
# adoption *type*, the *expected* activity counts, and z-score fallbacks, which
# come from app.growth_summary_mock (grep "# MOCK:") and vanish when
# GROWTH_SUMMARY_MOCK=false.

_SEX_KEY = {"male": "boys", "boys": "boys", "female": "girls", "girls": "girls"}


def _sex_key(gender: Optional[str]) -> Optional[str]:
    return _SEX_KEY.get((gender or "").strip().lower())


def _pick_bal(items: List[Any]) -> tuple:
    """Baseline / Assessment (mid) / Last from a chronological list."""
    if not items:
        return (None, None, None)
    return (items[0], items[len(items) // 2], items[-1])


def _activity_block(expected: Optional[int], actual: int) -> Dict[str, Any]:
    """Expected / Actual / Incomplete counts + percentages for one form group."""
    incomplete = max(expected - actual, 0) if expected is not None else None
    return {
        "expected": expected,
        "actual": actual,
        "incomplete": incomplete,
        "actual_pct": round(100 * actual / expected) if expected else None,
        "incomplete_pct": round(100 * incomplete / expected) if expected else None,
    }


def _visit_z(indicator: str, sex: Optional[str], visit: Optional[Dict[str, Any]]) -> Optional[float]:
    """One visit's z for an indicator.

    wfa: weight at age; lfa: length at age; wfl: weight at LENGTH (the WHO
    weight-for-height table is indexed by length-cm, not age).
    """
    if not visit or not sex:
        return None
    if indicator == "wfa":
        x, value = visit["age_days"], visit["weight"]
    elif indicator == "lfa":
        x, value = visit["age_days"], visit["length"]
    else:  # wfl
        x, value = visit["length"], visit["weight"]
    if x is None:
        return None
    return zscore_for_value(indicator, sex, x, value)


def _z_triplet(child_id: int, sex: Optional[str], visits: List[Dict[str, Any]],
               indicator: str) -> Dict[str, Optional[float]]:
    """WFAz/HFAz/WFHz at Baseline / Assessment / Last visit.

    A case with no measurement yet gets None, never a stand-in: a z-score reads
    as a clinical finding, and inventing one made the outcomes chart show bars
    for a child whose z-trend plots (which only ever plot real visits) correctly
    said "no measurements yet". Demo fall-backs are fine for programme config
    like expected counts; they are not fine for anthropometry.
    """
    bv, av, lv = _pick_bal(visits)
    return {
        label: _visit_z(indicator, sex, visit)
        for label, visit in (("bv", bv), ("av", av), ("lv", lv))
    }


def _summary_rows(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    today = date.today()
    for case in cases:
        child, mother, learner = case["child"], case["mother"], case["learner"]
        visits = case["visits"]
        child_id = child["id"]
        sex = _sex_key(child["gender"])

        # Adoption timing — REAL, from stored dates.
        dob = date.fromisoformat(child["dob"]) if child["dob"] else None
        adopted = date.fromisoformat(child["adoption_date"]) if child["adoption_date"] else None
        age_at_adoption = (adopted - dob).days if (dob and adopted) else None
        duration_days = (today - adopted).days if adopted else None
        timing_is_mock = False
        if gsm.MOCK_ENABLED and (age_at_adoption is None or duration_days is None):
            m_age, m_duration = gsm.mock_adoption_timing(child_id)  # MOCK: adoption timing
            age_at_adoption = m_age if age_at_adoption is None else age_at_adoption
            duration_days = m_duration if duration_days is None else duration_days
            timing_is_mock = True

        # Actual activity counts — REAL, one per (visit date) a form was filed.
        cg_actual = sum(1 for v in visits if GROWTH_FORM_KEY in v["forms"])
        bf_actual = sum(1 for v in visits if "breastfeeding" in v["forms"])
        cf_actual = sum(1 for v in visits if "complementary_feeding" in v["forms"])

        # Adoption type + expected counts — MOCK (null when disabled).
        adoption_type = None
        expected = {"cg": None, "bf": None, "cf": None}
        if gsm.MOCK_ENABLED:
            adoption_type = gsm.adoption_type(child_id, age_at_adoption)   # MOCK: adoption type
            expected = gsm.expected_counts(child_id, adoption_type, duration_days)  # MOCK: expected rules

        cg, bf, cf = (_activity_block(expected["cg"], cg_actual),
                      _activity_block(expected["bf"], bf_actual),
                      _activity_block(expected["cf"], cf_actual))
        total_expected = (sum(expected.values())
                          if all(v is not None for v in expected.values()) else None)
        total = _activity_block(total_expected, cg_actual + bf_actual + cf_actual)

        # Outcomes — REAL z-scores computed from measurements; null when there
        # are none. Never mock-filled (see _z_triplet).
        w_visits = [v for v in visits if v["weight"] and v["age_days"] is not None]
        l_visits = [v for v in visits if v["length"] and v["age_days"] is not None]
        wl_visits = [v for v in visits if v["weight"] and v["length"]]
        wfaz = _z_triplet(child_id, sex, w_visits, "wfa")
        hfaz = _z_triplet(child_id, sex, l_visits, "lfa")
        wfhz = _z_triplet(child_id, sex, wl_visits, "wfl")

        rows.append({
            "case_id": child_id,
            "learner_id": learner["id"],
            "identity": {
                "learner_name": learner["name"],
                "role_group": learner.get("role"),
                "mother_name": mother["name"],
                "child_uid": child["uid"],
            },
            "case_details": {
                "adoption_type": adoption_type,
                "age_of_adoption_days": age_at_adoption,
                "adoption_duration_days": duration_days,
            },
            "activities": {"total": total, "cg": cg, "bf": bf, "cf": cf},
            "outcomes": {"wfaz": wfaz, "hfaz": hfaz, "wfhz": wfhz},
            "meta": {
                "sex": sex,
                "district": learner.get("district"),
                "department": learner.get("department"),
                "expected_is_mock": bool(gsm.MOCK_ENABLED),
                "adoption_type_is_mock": bool(timing_is_mock),
                "timing_is_mock": bool(timing_is_mock),
                "z_is_mock": False,   # z-scores are always real or absent
            },
        })
    return rows


def _summary_filter_options(db: Session) -> Dict[str, Any]:
    """Distinct roles / departments / learners present among cases — feeds the
    growth-monitor filter dropdowns so they only offer values that return rows."""
    query = (
        db.query(
            models.User.id, models.User.full_name, models.User.email,
            models.User.role, models.User.department, models.ProgramDistrict.name,
        )
        .join(models.Mother, models.Mother.registered_by_user_id == models.User.id)
        .join(models.Child, models.Child.mother_id == models.Mother.id)
        .outerjoin(models.ProgramDistrict, models.User.program_district_id == models.ProgramDistrict.id)
        .distinct()
    )
    learners: Dict[int, Dict[str, Any]] = {}
    roles, departments = set(), set()
    for uid, full_name, email, role, department, district in query:
        department = gsm.resolved_department(uid, department)  # MOCK: department fill
        learners[uid] = {
            "id": uid, "name": full_name or email,
            "role": role, "department": department, "district": district,
        }
        if role:
            roles.add(role)
        if department:
            departments.add(department)
    return {
        "learners": sorted(learners.values(), key=lambda x: (x["name"] or "").lower()),
        "roles": sorted(roles),
        "departments": sorted(departments),
    }


# (group label, sub label, value getter, cell-fill getter) — mirrors the
# on-screen summary table, including its spectrum cell colours.
_ColSpec = Tuple[str, str, Callable[[Dict[str, Any]], Any], Optional[Callable[[Dict[str, Any]], Optional[str]]]]


def _months(days: Optional[int]) -> Optional[int]:
    return round(days / 30.44) if days is not None else None


# ── Excel-style spectrum fills (mirror frontend/src/lib/growthDisplay.ts) ────

# 7-stop ramp: deep red → red → orange → amber → lime → green → deep green.
# Keep in sync with SPECTRUM_STOPS in frontend/src/lib/growthDisplay.ts.
_SPECTRUM_STOPS = [
    (176, 0, 32), (226, 61, 40), (245, 124, 0), (245, 197, 24),
    (168, 201, 58), (76, 175, 80), (27, 127, 59),
]
_DEFAULT_ALPHA = 0.55


def _spectrum_hex(t: float, alpha: float = _DEFAULT_ALPHA) -> str:
    """Interpolate the ramp and flatten it over white by the same alpha the UI
    composites with, so the spreadsheet tint matches the on-screen cell."""
    t = max(0.0, min(1.0, t))
    scaled = t * (len(_SPECTRUM_STOPS) - 1)
    i = min(int(scaled), len(_SPECTRUM_STOPS) - 2)
    f = scaled - i
    lo, hi = _SPECTRUM_STOPS[i], _SPECTRUM_STOPS[i + 1]
    channels = (lo[c] + (hi[c] - lo[c]) * f for c in range(3))
    return "".join(f"{round(alpha * ch + (1 - alpha) * 255):02X}" for ch in channels)


def _z_fill(z: Optional[float]) -> Optional[str]:
    """Deep red at ≤ −3, red around −2, through amber, green at 0 and above."""
    if z is None:
        return None
    # ±3 SD and beyond is painted near-solid so the flag is unmistakable
    # (mirrors zAlpha in frontend/src/lib/growthDisplay.ts).
    return _spectrum_hex((z + 3) / 3, 0.92 if abs(z) >= 3 else _DEFAULT_ALPHA)


def _pct_fill(pct: Optional[float], invert: bool = False) -> Optional[str]:
    """AC: 0 % of expected = red → 100 %+ = green. IC (invert): 0 % = green."""
    if pct is None:
        return None
    return _spectrum_hex(((100 - pct) if invert else pct) / 100)


def _xlsx_columns() -> List[_ColSpec]:
    def activity(group: str, key: str) -> List[_ColSpec]:
        def ac_fill(r: Dict[str, Any], k: str = key) -> Optional[str]:
            return _pct_fill(r["activities"][k]["actual_pct"])

        def ic_fill(r: Dict[str, Any], k: str = key) -> Optional[str]:
            return _pct_fill(r["activities"][k]["incomplete_pct"], invert=True)

        return [
            (group, "EC", lambda r, k=key: r["activities"][k]["expected"], None),
            (group, "AC", lambda r, k=key: r["activities"][k]["actual"], ac_fill),
            (group, "AC %", lambda r, k=key: r["activities"][k]["actual_pct"], ac_fill),
            (group, "IC", lambda r, k=key: r["activities"][k]["incomplete"], ic_fill),
            (group, "IC %", lambda r, k=key: r["activities"][k]["incomplete_pct"], ic_fill),
        ]

    def zcol(sub: str, metric: str, visit: str) -> _ColSpec:
        return (
            "Outcomes (z)",
            sub,
            lambda r, m=metric, v=visit: r["outcomes"][m][v],
            lambda r, m=metric, v=visit: _z_fill(r["outcomes"][m][v]),
        )

    return [
        ("Identity", "Learner", lambda r: r["identity"]["learner_name"], None),
        ("Identity", "Role", lambda r: r["identity"]["role_group"], None),
        ("Identity", "Mother", lambda r: r["identity"]["mother_name"], None),
        ("Identity", "Child ID", lambda r: r["identity"]["child_uid"], None),
        ("Case details", "Adoption type", lambda r: r["case_details"]["adoption_type"], None),
        ("Case details", "Adopt age (mo)", lambda r: _months(r["case_details"]["age_of_adoption_days"]), None),
        ("Case details", "Duration (mo)", lambda r: _months(r["case_details"]["adoption_duration_days"]), None),
        *activity("Total activities", "total"),
        # Outcomes sit between Total and the per-form groups (client request);
        # headers put the visit code before the z (W·BVz), and weight-for-height
        # (WH) completes the WHO triad.
        zcol("W·BVz", "wfaz", "bv"),
        zcol("W·AVz", "wfaz", "av"),
        zcol("W·LVz", "wfaz", "lv"),
        zcol("H·BVz", "hfaz", "bv"),
        zcol("H·AVz", "hfaz", "av"),
        zcol("H·LVz", "hfaz", "lv"),
        zcol("WH·BVz", "wfhz", "bv"),
        zcol("WH·AVz", "wfhz", "av"),
        zcol("WH·LVz", "wfhz", "lv"),
        *activity("CG (Check Growth)", "cg"),
        *activity("BF (Breastfeeding)", "bf"),
        *activity("CF (Compl. Feeding)", "cf"),
    ]


def _summary_xlsx(rows: List[Dict[str, Any]]) -> BytesIO:
    """Render the summary rows to an .xlsx with a merged two-row grouped header,
    frozen identity block + header, and an auto-filter — mirroring the UI table."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = _xlsx_columns()
    wb = Workbook()
    ws = wb.active
    ws.title = "Growth Summary"

    bold = Font(bold=True)
    group_fill = PatternFill("solid", fgColor="E8E0D8")
    sub_fill = PatternFill("solid", fgColor="F4EFE9")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 2: per-column sub-headers.
    for ci, (_group, sub, _getter, _fill) in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=sub)
        cell.font = bold
        cell.fill = sub_fill
        cell.alignment = center

    # Row 1: group headers, merged across their sub-columns.
    ci = 1
    while ci <= len(cols):
        group = cols[ci - 1][0]
        span = 1
        while ci + span <= len(cols) and cols[ci - 1 + span][0] == group:
            span += 1
        cell = ws.cell(row=1, column=ci, value=group)
        cell.font = bold
        cell.fill = group_fill
        cell.alignment = center
        if span > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + span - 1)
        ci += span

    # Data rows — value cells carry the same spectrum fills as the UI table.
    for ri, row in enumerate(rows, start=3):
        for ci, (_group, _sub, getter, fill) in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci, value=getter(row))
            color = fill(row) if fill else None
            if color:
                cell.fill = PatternFill("solid", fgColor=color)

    for ci, (_group, sub, _getter, _fill) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = max(9, min(22, len(sub) + 3))
    ws.freeze_panes = "E3"  # freeze the 4 Identity columns + the two header rows
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(rows) + 2}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/standards")
def get_standards():
    """WHO percentile curves for all indicators/sexes — public reference data
    (published WHO tables, nothing user-specific); the frontend caches this."""
    return percentile_curves()


@router.get("/my-cases")
def my_growth_cases(
    current_user: models.User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    return {"cases": _build_cases(db, user_id=current_user.id)}


@admin_router.get("/monitor")
def admin_growth_monitor(
    district: str = Query("", description="Program-district slug; empty = all districts"),
    role: str = Query("", description="Learner role/designation; empty = all"),
    department: str = Query("", description="Learner department; empty = all"),
    learner_id: Optional[int] = Query(None, description="Single learner id; null = all"),
    child_id: Optional[int] = Query(None, description="Single case (child) id; null = all"),
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    return {"cases": _build_cases(
        db, district_slug=district or None, role=role or None,
        department=department or None, learner_id=learner_id, child_id=child_id,
    )}


@admin_router.get("/summary")
def admin_growth_summary(
    district: str = Query(""),
    role: str = Query(""),
    department: str = Query(""),
    learner_id: Optional[int] = Query(None),
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Learner-level summary table: EC/AC/IC per form group + WFAz/HFAz outcomes,
    one row per learner-mother-child case, honouring the four filters."""
    cases = _build_cases(
        db, district_slug=district or None, role=role or None,
        department=department or None, learner_id=learner_id,
    )
    return {"rows": _summary_rows(cases), "mock": bool(gsm.MOCK_ENABLED)}


@admin_router.get("/alerts")
def admin_growth_alerts(
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Submitted protein assessments whose 24-hour total was flagged as
    implausibly high (see forms.PROTEIN_HIGH_TOTAL_G).

    The learner and every admin also get a notification at submit time; this is
    the durable list the team can work through, newest first.
    """
    responses = (
        db.query(models.FormResponse)
        .filter(
            models.FormResponse.form_key == "mother_protein_intake",
            models.FormResponse.status == "submitted",
        )
        .order_by(models.FormResponse.assessment_date.desc(), models.FormResponse.id.desc())
        .limit(500)
        .all()
    )
    alerts: List[Dict[str, Any]] = []
    for r in responses:
        protein = ((r.summary_json or {}).get("protein") or {})
        if not protein.get("high24"):
            continue
        mother = db.query(models.Mother).filter(models.Mother.id == r.mother_id).first()
        learner = (
            db.query(models.User).filter(models.User.id == mother.registered_by_user_id).first()
            if mother is not None else None
        )
        alerts.append({
            "response_id": r.id,
            "assessment_date": r.assessment_date.isoformat() if r.assessment_date else None,
            "mother_id": r.mother_id,
            "mother_name": mother.mother_name if mother else None,
            "learner_name": (learner.full_name or learner.email) if learner else None,
            "total24": protein.get("total24"),
        })
        if len(alerts) >= limit:
            break
    return {"alerts": alerts, "threshold": PROTEIN_HIGH_TOTAL_G}


@admin_router.get("/summary/filters")
def admin_growth_summary_filters(
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    return _summary_filter_options(db)


@admin_router.get("/summary/export")
def admin_growth_summary_export(
    district: str = Query(""),
    role: str = Query(""),
    department: str = Query(""),
    learner_id: Optional[int] = Query(None),
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """The (filtered) summary table as an .xlsx download."""
    cases = _build_cases(
        db, district_slug=district or None, role=role or None,
        department=department or None, learner_id=learner_id,
    )
    buffer = _summary_xlsx(_summary_rows(cases))
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="growth-summary.xlsx"'},
    )


@admin_router.get("/cases/{child_id}/detail")
def admin_growth_case_detail(
    child_id: int,
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Full drill-down for one learner-mother-child case: identity plus EVERY
    submitted form — child-level responses grouped by visit date, and mother-level
    forms (antenatal, protein intake) — each with its complete answer detail."""
    child = db.query(models.Child).filter(models.Child.id == child_id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Child not found")
    mother = db.query(models.Mother).filter(models.Mother.id == child.mother_id).first()
    learner = None
    if mother and mother.registered_by_user_id:
        learner = db.query(models.User).filter(models.User.id == mother.registered_by_user_id).first()

    child_responses = (
        db.query(models.FormResponse)
        .filter(models.FormResponse.child_id == child_id, models.FormResponse.status == "submitted")
        .order_by(models.FormResponse.assessment_date, models.FormResponse.id)
        .all()
    )
    by_date: Dict[Any, List[models.FormResponse]] = defaultdict(list)
    for response in child_responses:
        by_date[response.assessment_date].append(response)

    dob = child.dob
    sex = _sex_key(child.gender)
    visits: List[Dict[str, Any]] = []
    for visit_date in sorted(k for k in by_date if k is not None):
        responses = by_date[visit_date]
        weight = length = None
        for response in responses:
            if response.form_key == GROWTH_FORM_KEY:
                metrics = _extract_metrics(response.answers_json)
                weight, length = metrics["weight"], metrics["length"]
        visits.append({
            "date": visit_date.isoformat(),
            "age_days": (visit_date - dob).days if dob else None,
            "weight": weight,
            "length": length,
            "responses": [_serialize_detail(r, child, mother) for r in responses],
        })

    # Per-visit z (this visit's own measurements) and cumulative BV/AV/LV
    # triplets over the visits up to and including each one. REAL values only —
    # the drill-down never mock-fills; missing measurements stay null.
    for i, visit in enumerate(visits):
        visit["z"] = {
            "wfaz": _visit_z("wfa", sex, visit if visit["weight"] and visit["age_days"] is not None else None),
            "hfaz": _visit_z("lfa", sex, visit if visit["length"] and visit["age_days"] is not None else None),
            "wfhz": _visit_z("wfl", sex, visit if visit["weight"] and visit["length"] else None),
        }
        upto = visits[: i + 1]
        to_date: Dict[str, Any] = {}
        for out_key, indicator, keep in (
            ("wfaz", "wfa", lambda v: v["weight"] and v["age_days"] is not None),
            ("hfaz", "lfa", lambda v: v["length"] and v["age_days"] is not None),
            ("wfhz", "wfl", lambda v: v["weight"] and v["length"]),
        ):
            measured = [v for v in upto if keep(v)]
            bv, av, lv = _pick_bal(measured)
            to_date[out_key] = {
                "bv": _visit_z(indicator, sex, bv),
                "av": _visit_z(indicator, sex, av),
                "lv": _visit_z(indicator, sex, lv),
            }
        visit["z_to_date"] = to_date

    mother_forms: List[Dict[str, Any]] = []
    if mother:
        mother_responses = (
            db.query(models.FormResponse)
            .filter(
                models.FormResponse.mother_id == mother.id,
                models.FormResponse.child_id.is_(None),
                models.FormResponse.status == "submitted",
            )
            .order_by(models.FormResponse.assessment_date, models.FormResponse.id)
            .all()
        )
        mother_forms = [_serialize_detail(r, None, mother) for r in mother_responses]

    return {
        "child": {
            "id": child.id, "uid": child.child_uid, "name": child.child_name,
            "gender": child.gender,
            "dob": dob.isoformat() if dob else None,
            "birth_weight": child.birth_weight, "birth_length": child.birth_length,
            "adoption_date": child.adoption_date.isoformat() if child.adoption_date else None,
        },
        "mother": {
            "id": mother.id, "uid": mother.mother_uid, "name": mother.mother_name,
            "mobile": mother.mobile, "village": mother.village, "age": mother.mother_age,
        } if mother else None,
        "learner": {
            "id": learner.id,
            "name": learner.full_name or learner.email,
            "role": learner.role,
            "department": gsm.resolved_department(learner.id, learner.department),  # MOCK: department fill
            "email": learner.email,
        } if learner else None,
        "visits": visits,
        "mother_forms": mother_forms,
    }


@admin_router.get("/responses/{response_id}")
def admin_response_detail(
    response_id: int,
    admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    response = (
        db.query(models.FormResponse)
        .filter(models.FormResponse.id == response_id)
        .first()
    )
    if not response:
        raise HTTPException(status_code=404, detail="Response not found")
    child = db.query(models.Child).filter(models.Child.id == response.child_id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Child not found")
    return _serialize_detail(response, child)
